#read an excel workbook and translate each sheet into LaTeX code
from openpyxl import load_workbook
from sys import argv

wb=load_workbook(filename=argv[1], data_only=True)

prec=3
if len(argv)>2:
    prec=int(argv[2])

#remove zero before the decimal point in floats
removezero=False
#use \resizebox to fit the table to the column in latex
fillcolumn=False

for sheet in wb.worksheets:
    print("\\begin{table}[h]")
    print("  \\begin{center}")
    print("    \\caption{" + wb.sheetnames[wb.index(sheet)]+ "}")
    print("    \\label{" + "".join(wb.sheetnames[wb.index(sheet)].split(" ")) + "}")
    if fillcolumn:
        print("    \\resizebox{\\columnwidth}{!}{")
    print("    \\begin{tabular}{"+"r|"*len(next(sheet.rows))+"\b}")
    for row in sheet.rows:
        print("      ", end="")
        for x in row:
            val = x.value
            if val==None:
                val=""
            elif type(val)==type(1.0):
                val = str(round(val, prec))
                if removezero and val[0]=="0":
                    val=val[1:]
            elif type(val)==type("s"):
                val=val.replace("&", "\\&").replace("%", "\\%")

            print(val, end=" & ")
        print("\b\b\\\\")
    print("    \\end{tabular}")
    if fillcolumn:
        print("    }")
    print("  \\end{center}")
    print("\\end{table}")
    print("\n")
    
